from datetime import date
from tracker import SleepTracker
from actracker import activity_tracker
import matplotlib.pyplot as plt
import openpyxl as pyxl
from openweather_ import get_weat
import tkinter
import customtkinter
from tkinter import *
# name age date hours quality dream
# name=input("What's your name? ")
# age=input("How old are you? ")
# hours=input("How long have you slept?(in hours) ")
# quality=input("How did you sleep sleep (Good, Bad or Mediocre) ")
# dream=input("Did you have a dream (Yes or No) ")
# st=SleepTracker(name,age,date.today(),hours,quality,dream,get_weat())
# e_hours=input("How long you have spent on the activity?")
# what_did=input("What did you do?")
# feelings=input("How did you feel?")
# at=activity_tracker(e_hours,what_did,feelings)
# at.analyze_activity()
# st.data_store()
# at.data_store()
# wb=pyxl.load_workbook("data.xlsx")
# sheet=wb.active
# sl=[]
# dt=[]
# for i in range(1,sheet.max_row+1):
#     sl.append(sheet["D"+str(i)].value)
#     dt.append(i)
# plt.plot(dt,sl,label="sleep hours")

# sl=[]
# dt=[]
# for i in range(1,sheet.max_row+1):
#     sl.append(sheet["G"+str(i)].value)
#     dt.append(i)
# plt.plot(dt,sl,label="activity hours")
# plt.title("activity hours and sleep hours")
# plt.legend()
# plt.show()    

# # import matplotlib
# # import openpyxl as pyxl
# # wb=pyxl.load_workbook("data.xlsx")
# # sheet=wb.active
# # for i in range(1,11):
# #     index="A"+str(i)
# #     sheet[index]="hello"
# #     print(sheet[index].value)
# # wb.save("data.xlsx")


global entry1, entry2, entry3, entry4, entry5, entry6, entry7, entry8
root=customtkinter.CTk()
root.title("trackerapp")
root.geometry("600x500")

def analyze():
    st=SleepTracker(entry1.get(),int(entry2.get()),date.today(),entry3.get(),entry4.get(),entry5.get(),get_weat())
    st.data_store()
    wb=pyxl.load_workbook("data.xlsx")
    sheet=wb.active
    sl=[]
    dt=[]
    for i in range(1,sheet.max_row+1):
        sl.append(sheet["D"+str(i)].value)
        dt.append(i)
    plt.plot(dt,sl,label="sleep hours")
    plt.show()

frame=customtkinter.CTkFrame(master=root)
frame.pack(pady=20,padx=10)
entry1=customtkinter.CTkEntry(master=frame,placeholder_text="What's your name? ",width=300,justify=CENTER)
entry1.pack(pady=12,padx=10)

entry2=customtkinter.CTkEntry(master=frame,placeholder_text="How old are you?",width=300,justify=CENTER)
entry2.pack(pady=12,padx=10)

entry3=customtkinter.CTkEntry(master=frame,placeholder_text="How long have you slept?(in hours)",width=300,justify=CENTER)
entry3.pack(pady=12,padx=10)

entry4=customtkinter.CTkEntry(master=frame,placeholder_text="How did you sleep sleep? (Good, Bad or Mediocre)",width=400,justify=CENTER)
entry4.pack(pady=12,padx=10)

entry5=customtkinter.CTkEntry(master=frame,placeholder_text="Did you have a dream? (Yes or No)",width=300,justify=CENTER)
entry5.pack(pady=12,padx=10)

entry6=customtkinter.CTkEntry(master=frame,placeholder_text="How long you have spent on the activity?",width=300,justify=CENTER)
entry6.pack(pady=12,padx=10)

entry7=customtkinter.CTkEntry(master=frame,placeholder_text="What did you do?",width=300,justify=CENTER)
entry7.pack(pady=12,padx=10)

entry8=customtkinter.CTkEntry(master=frame,placeholder_text="How did you feel?",width=300,justify=CENTER)
entry8.pack(pady=12,padx=10)

button=customtkinter.CTkButton(master=frame,text="analyze",command=analyze).pack(pady=12,padx=10)

root.mainloop()